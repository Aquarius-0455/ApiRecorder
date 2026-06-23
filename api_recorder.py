import os
import sys
os.environ.setdefault("NODE_NO_WARNINGS", "1")  # 隐藏 Playwright 内部 Node.js 的 DeprecationWarning

import asyncio
import json
import re
import yaml
from urllib.parse import urlparse, parse_qs
from playwright.async_api import async_playwright

# ==========================================
# 环境与启动配置
# ==========================================
class BrowserConfig:
    """环境感知配置类"""
    ENV = os.getenv("RECORDER_ENV", "local")
    CDP_URL = os.getenv("CDP_URL", None)
    IS_REMOTE = os.getenv("IS_REMOTE", "false").lower() == "true"
    
    # 默认浏览器启动参数
    LAUNCH_KWARGS = {
        "headless": False,
        "args": ["--start-maximized", "--window-size=1920,1080"] if ENV == "local" else []
    }
    
    # 默认上下文参数
    CONTEXT_KWARGS = {
        "no_viewport": True,
        "ignore_https_errors": True
    }

class RecorderSettings:
    """录制器全局设置开关"""
    BEAUTIFY_YAML = True
    NORMALIZE_PATH = True
    REDACT_TOKEN = True  # 是否对 Authorization Token 进行脱敏

class APIRecorder:
    def __init__(self, module_name):
        self.module_name = module_name
        self.captured_apis = {}
        self.is_recording = False
        self.ignore_patterns = [
            r'\.js', r'\.css', r'\.png', r'\.jpg', r'\.gif', r'\.svg', r'\.woff',
            r'google-analytics', r'sentry.io', r'hotjar', r'log-upload'
        ]
        # 移除 keep_headers 白名单，改用简单的“外部设置黑名单”
        self.exclude_headers = ['host', 'connection', 'content-length', 'expect']

    def is_target_url(self, url):
        for pattern in self.ignore_patterns:
            if re.search(pattern, url, re.I):
                return False
        return True

    def normalize_path(self, url):
        parsed = urlparse(url)
        path = parsed.path
        segments = path.split('/')
        normalized_segments = []
        path_params = []
        param_index = 1
        
        for i, seg in enumerate(segments):
            if not seg:
                normalized_segments.append("")
                continue
            if re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$|^\d{5,}$|^[A-Z0-9]{8,}$', seg):
                prev_seg = segments[i-1] if i > 0 else "id"
                base_name = prev_seg.rstrip('s') if prev_seg.endswith('s') else prev_seg
                param_name = f"{base_name}_id"
                if param_name in path_params:
                    param_name = f"{param_name}_{param_index}"
                    param_index += 1
                normalized_segments.append(f"{{{param_name}}}")
                path_params.append(param_name)
            else:
                normalized_segments.append(seg)
        return "/".join(normalized_segments), path_params

    async def handle_response(self, response):
        if not self.is_recording:
            return
        request = response.request
        if request.resource_type in ["fetch", "xhr"] and self.is_target_url(request.url):
            try:
                url, method = request.url, request.method
                headers = await request.all_headers()
                
                # 记录所有 Header（排除掉不能手动设置的协议头）
                safe_headers = {}
                for k, v in headers.items():
                    k_low = k.lower()
                    if k_low in self.exclude_headers or k_low.startswith(':'):
                        continue
                    
                    # 依然保留 Token 脱敏逻辑
                    if k_low == 'authorization' and RecorderSettings.REDACT_TOKEN:
                        if v.lower().startswith('bearer '):
                            safe_headers[k_low] = "Bearer TOKEN"
                        else:
                            safe_headers[k_low] = "REDACTED"
                    else:
                        safe_headers[k_low] = v
                
                response_body = None
                try: response_body = await response.json()
                except:
                    try: response_body = await response.text()
                    except: response_body = "<Binary Data>"

                post_data = request.post_data
                body, body_type = None, "json"
                if post_data:
                    try: body, body_type = json.loads(post_data), "json"
                    except: body, body_type = post_data, "raw"
                elif "?" in url:
                    body_type = "params"
                    query = urlparse(url).query
                    if query:
                        raw = {k: v[0] if len(v)==1 else v for k, v in parse_qs(query).items()}
                        body = {k: self._infer_type(v) for k, v in raw.items()} if RecorderSettings.BEAUTIFY_YAML else raw

                # 计算归一化路径（提前，用于去重与 ID 分配）
                norm_path, path_params = self.normalize_path(url) if RecorderSettings.NORMALIZE_PATH else (urlparse(url).path, [])

                # 去重逻辑: 归一化路径 + 方法 + body 完全一致 → 视为重复，忽略
                for item in self.captured_apis.values():
                    if (item['method'] == method
                            and item.get('norm_path') == norm_path
                            and item['body'] == body):
                        return

                # 为相同路径但不同参数的接口分配唯一 ID
                base_id = f"{method}_{norm_path}"
                api_id = base_id
                idx = 1
                while api_id in self.captured_apis:
                    api_id = f"{base_id}_{idx}"
                    idx += 1

                self.captured_apis[api_id] = {
                    "method": method, "url": url, "norm_path": norm_path,
                    "path_params": path_params, "headers": safe_headers,
                    "body": body, "body_type": body_type, "response_body": response_body, "status": response.status
                }
                print(f"  [+] Recorded: {method} {norm_path} (ID: {api_id})")
            except: pass

    def _infer_type(self, val):
        if isinstance(val, list): return [self._infer_type(item) for item in val]
        if not isinstance(val, str): return val
        low = val.lower()
        if low == 'true': return True
        if low == 'false': return False
        if low == 'null': return None
        try: return float(val) if '.' in val else int(val)
        except: return val

    def _format_yaml_value(self, value):
        if isinstance(value, bool): return "true" if value else "false"
        if isinstance(value, (int, float)): return str(value)
        if value is None: return "null"
        if value == "": return '""'
        escaped = str(value).replace('\\', '\\\\').replace('"', '\\"')
        return f'"{escaped}"'

    def _get_comment(self, value):
        if not RecorderSettings.BEAUTIFY_YAML: return ""
        if isinstance(value, bool): return "# 布尔值"
        if isinstance(value, int): return "# 整数"
        if isinstance(value, float): return "# 浮点数"
        if isinstance(value, str): return "# 字符串"
        if value is None: return "# 空值"
        return ""

    def _dict_to_yaml_lines(self, d, indent=0):
        lines = []
        prefix = " " * indent
        max_base_len = 0
        if RecorderSettings.BEAUTIFY_YAML and d:
            for key, value in d.items():
                if not isinstance(value, (dict, list)):
                    base_line = f"{prefix}{key}: {self._format_yaml_value(value)}"
                    max_base_len = max(max_base_len, len(base_line))

        for key, value in d.items():
            if isinstance(value, dict):
                lines.append(f"{prefix}{key}:")
                lines.extend(self._dict_to_yaml_lines(value, indent + 2))
            elif isinstance(value, list):
                if not value:
                    lines.append(f"{prefix}{key}: []")
                elif all(isinstance(item, (str, int, float, bool)) for item in value):
                    # 简单列表，用行内格式
                    items_str = ", ".join(self._format_yaml_value(i) for i in value)
                    lines.append(f"{prefix}{key}: [{items_str}]")
                else:
                    # 复杂列表（含嵌套对象）
                    lines.append(f"{prefix}{key}:")
                    for item in value:
                        if isinstance(item, dict):
                            first = True
                            for k2, v2 in item.items():
                                marker = "- " if first else "  "
                                if isinstance(v2, dict):
                                    lines.append(f"{prefix}  {marker}{k2}:")
                                    lines.extend(self._dict_to_yaml_lines(v2, indent + 6))
                                else:
                                    lines.append(f"{prefix}  {marker}{k2}: {self._format_yaml_value(v2)}")
                                first = False
                        else:
                            lines.append(f"{prefix}  - {self._format_yaml_value(item)}")
            else:
                formatted = self._format_yaml_value(value)
                line = f"{prefix}{key}: {formatted}"
                if RecorderSettings.BEAUTIFY_YAML:
                    comment = self._get_comment(value)
                    if comment:
                        padding = " " * max(2, (max_base_len + 2) - len(line))
                        line += f"{padding}{comment}"
                lines.append(line)
        return lines

    def generate_yaml(self):
        y = [f'module: "{self.module_name}"', "", "apis:"]
        for api_id, data in self.captured_apis.items():
            method = data['method'].upper()
            # 优先使用录制时缓存的归一化路径，回退到重新计算
            path = data.get('norm_path') or (
                self.normalize_path(data['url'])[0] if RecorderSettings.NORMALIZE_PATH else urlparse(data['url']).path
            )
            params = data.get('path_params') or re.findall(r'\{(\w+)\}', path)
            
            has_id = '{' in path
            if method == "POST" and not has_id: key = "save"
            elif method == "GET" and not has_id: key = "list"
            elif method == "GET" and has_id: key = "detail"
            elif method in ["PUT", "PATCH"]: key = "update"
            elif method == "DELETE": key = "delete"
            else:
                parts = [p for p in path.split('/') if p and '{' not in p]
                key = f"{method.lower()}_{parts[-1].replace('-', '_')}" if parts else "root"
            
            idx = 1
            final_key = key
            while any(line.strip().startswith(f"{final_key}:") for line in y):
                final_key = f"{key}_{idx}"; idx += 1

            y.append(f"  # {method} {path}")
            y.append(f"  {final_key}:")
            y.append(f'    name: "{method} {path}"\n    path: "{path}"\n    method: "{method}"')
            if data['headers']:
                y.append("    headers:")
                for k, v in data['headers'].items(): y.append(f'      {k}: {self._format_yaml_value(v)}')
            y.append(f'    body_type: "{data["body_type"]}"')
            if data['body']:
                if RecorderSettings.BEAUTIFY_YAML:
                    # 展开格式：多行 YAML，带类型推导和注释
                    y.append("    default_body:")
                    y.extend(self._dict_to_yaml_lines(data['body'], 6))
                else:
                    # 紧凑格式：行内 JSON {}，原汁原味
                    y.append(f"    default_body: {json.dumps(data['body'], ensure_ascii=False)}")
            if params:
                y.append("    path_params:")
                for p in params: y.append(f'      - {p}')
            y.append('    assertions:\n      - type: "status_code"\n        expected: 200\n      - type: "json_path"\n        path: "code"\n        expected: 200\n')
        return "\n".join(y)

    async def run(self):
        async with async_playwright() as p:
            print("\n" + "═"*55)
            print("  ⚡  API Traffic Recorder  |  流量录制工具")
            print("═"*55)
            print("🔍 正在检查 9222 端口是否有可复用的浏览器会话...")
            try:
                browser = await p.chromium.connect_over_cdp("http://localhost:9222", timeout=3000)
                print("✅ 已复用已有浏览器")
            except:
                print("💻 未找到已有会话，正在启动全新浏览器...")
                args = BrowserConfig.LAUNCH_KWARGS.copy()
                args.setdefault("args", []).append("--remote-debugging-port=9222")
                browser = await p.chromium.launch(**args)

            context = browser.contexts[0] if browser.contexts else await browser.new_context(**BrowserConfig.CONTEXT_KWARGS)
            
            # [核心修复] 监听后续产生的所有新页面 (解决新窗口/新标签页流量无法抓取的问题)
            context.on("page", lambda p: p.on("response", self.handle_response))
            
            # [核心修复] 绑定当前已有的所有页面
            for p in context.pages:
                p.on("response", self.handle_response)

            # 获取主页面用于跳转落地页（若已打开页面则复用，否则新建）
            page = context.pages[0] if context.pages else await context.new_page()
            
            index_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")
            if os.path.exists(index_path):
                try:
                    await page.goto("file:///" + index_path.replace("\\", "/"), wait_until="domcontentloaded", timeout=5000)
                    print("🌐 落地页已加载")
                except:
                    pass

            print("\n" + "─"*55)
            print(f" 📦 模块名称 : {self.module_name}")
            print("─"*55)
            print("  命令速查:")
            print("    s  → 开始抓取")
            print("    p  → 暂停抓取")
            print("    c  → 清空已抓取数据")
            print("    n  → 切换路径归一化")
            print("    b  → 切换YAML美化推导")
            print("    t  → 切换Token自动脱敏")
            print("    e  → 导出并退出")
            print("─"*55 + "\n")

            while True:
                count = len(self.captured_apis)
                status = "🔴 录制中" if self.is_recording else "⏹️ 已暂停"
                norm = "开" if RecorderSettings.NORMALIZE_PATH else "关"
                beautify = "开" if RecorderSettings.BEAUTIFY_YAML else "关"
                redact = "开" if RecorderSettings.REDACT_TOKEN else "关"
                
                cmd_line = await asyncio.to_thread(
                    input,
                    f"[{status}] \n 已抓 {count} 条 \n 归一化:{norm} | 美 化:{beautify} | 脱 敏:{redact} \n\n 命令(s/p/c/n/b/t/e):\n "
                )
                cmd_line = cmd_line.strip().lower()

                if not cmd_line: continue

                # 支持组合命令，如 "nbt"
                should_exit = False
                for cmd in cmd_line:
                    if cmd == 's':
                        self.is_recording = True
                        print(" ▶️ 抓取已开启\n")
                    elif cmd == 'p':
                        self.is_recording = False
                        print(" ⏹️ 抓取已暂停\n")
                    elif cmd == 'c':
                        self.captured_apis.clear()
                        print(" 🧹 已抓取的数据已清空，可以重新录制\n")
                    elif cmd == 'n':
                        RecorderSettings.NORMALIZE_PATH = not RecorderSettings.NORMALIZE_PATH
                        state = "开启" if RecorderSettings.NORMALIZE_PATH else "关闭"
                        print(f" 🔄 路径归一化已{state}\n")
                    elif cmd == 'b':
                        RecorderSettings.BEAUTIFY_YAML = not RecorderSettings.BEAUTIFY_YAML
                        state = "开启" if RecorderSettings.BEAUTIFY_YAML else "关闭"
                        tip = "（数字/布尔类型自动推导）" if RecorderSettings.BEAUTIFY_YAML else "（保留原始字符串）"
                        print(f" 🎨 YAML 美化推导已{state} {tip}\n")
                    elif cmd == 't':
                        RecorderSettings.REDACT_TOKEN = not RecorderSettings.REDACT_TOKEN
                        state = "开启" if RecorderSettings.REDACT_TOKEN else "关闭"
                        tip = "（Authorization 自动替换）" if RecorderSettings.REDACT_TOKEN else "（保留原始 Token）"
                        print(f" 🛡️ Token 自动脱敏已{state} {tip}\n")
                    elif cmd == 'e':
                        actual_count = len(self.captured_apis)
                        print("\n" + "─"*55)
                        print(f" 💾 正在导出，共 {actual_count} 条接口...")
                        self.save_to_file()
                        print(" ✅ 导出完成！文件保存在 captured_apis/ 目录下")
                        print("─"*55 + "\n")
                        should_exit = True
                        break
                    elif cmd_line == 'exit':
                        print("  👋 直接退出，未保存。")
                        should_exit = True
                        break
                    else:
                        if len(cmd_line) == 1: # 只有单字符报错时才提示，避免多字符混合时刷屏
                            print(f" ❓ 未知命令: {cmd}\n")
                
                if should_exit: break


    def save_to_file(self):
        if not self.captured_apis: return
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "captured_apis")
        os.makedirs(path, exist_ok=True)
        with open(os.path.join(path, f"{self.module_name}.yaml"), 'w', encoding='utf-8') as f:
            f.write(self.generate_yaml())
        self._save_excel(os.path.join(path, f"{self.module_name}.xlsx"))

    def _save_excel(self, file_path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "API Data"
            ws.freeze_panes = "A2"
            
            # 样式定义
            header_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
            header_font = Font(bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # 表头
            headers = ["ID", "Name", "Method", "Path", "Headers", "Body Type", "Default Body", "Response Body", "Assertions"]
            ws.append(headers)
            ws.row_dimensions[1].height = 30
            
            # 应用表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 写入数据
            for idx, (api_id, data) in enumerate(self.captured_apis.items(), 1):
                norm_path, _ = self.normalize_path(data['url']) if RecorderSettings.NORMALIZE_PATH else (urlparse(data['url']).path, [])
                
                def try_parse_json_strings(obj):
                    if isinstance(obj, dict):
                        return {k: try_parse_json_strings(v) for k, v in obj.items()}
                    elif isinstance(obj, list):
                        return [try_parse_json_strings(i) for i in obj]
                    elif isinstance(obj, str):
                        s = obj.strip()
                        if (s.startswith('{') and s.endswith('}')) or (s.startswith('[') and s.endswith(']')):
                            try:
                                return try_parse_json_strings(json.loads(s))
                            except:
                                pass
                    return obj

                def safe_json(val):
                    if not val: return ""
                    parsed_val = try_parse_json_strings(val)
                    if isinstance(parsed_val, (dict, list)):
                        return json.dumps(parsed_val, indent=2, ensure_ascii=False)
                    return str(parsed_val)

                row = [
                    idx, # 使用自增 ID
                    f"{data['method']} {norm_path}",
                    data['method'],
                    norm_path,
                    safe_json(data['headers']),
                    data['body_type'],
                    safe_json(data['body']),
                    safe_json(data['response_body']),
                    "status: 200"
                ]
                ws.append(row)
                
                # 应用行样式
                for cell in ws[ws.max_row]:
                    cell.alignment = left_alignment
                    cell.border = thin_border
            
            # 设置列宽
            col_widths = {'A': 20, 'B': 30, 'C': 10, 'D': 40, 'E': 30, 'F': 12, 'G': 40, 'H': 50, 'I': 20}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(file_path)
            print(f"✅ Excel Exported: {file_path}")
        except Exception as e:
            print(f"⚠️ Excel skipped: {str(e)}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--module", default="captured_module")
    args = parser.parse_args()
    asyncio.run(APIRecorder(args.module).run())
